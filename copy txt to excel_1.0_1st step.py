import openpyxl
import os

# show informations
print("This is version 1.0 1st step")

# create a workbook
wb = openpyxl.Workbook()
wb.guess_types = True # avoid number stored in string style

# txt file number
num = 0

# get the path of target folder
path = input("The path of target folder:")
folder_name = path.split("\\")[-1] # split with "\" and select the last one
print("folder name: " + str(folder_name))

# If there is txt, read it and paste it into a new sheet
for file in os.listdir(path):
    if file.endswith(".txt"):
        print(str(num) + ". " + file)
        with open(path + "\\" + file, "r") as file_obj:
            file_name = file_obj.name.split('\\')[-1]
            name = file_name.rsplit('.',1)[0] # .rsplit('.',1)[0] to remove ".txt"
            wb.create_sheet(title = name) # create worksheet
            ws = wb[name] # select the worksheet we created
            
            log_list = file_obj.readlines()
            length = len(log_list)
            print("cell number: " + str(length-1))
            
            for i in range(0,length):
                log_list_withoutn = log_list[i].split("\n")[0] # remove \n
                ws.append(log_list_withoutn.split("\t"))

            ws["I1"] = "average"
            ws["J1"] = "SD"
            ws["K1"] = "SEM"

            ws["I2"] = "=AVERAGE(H2:H" + str(length) + ")"
            ws["J2"] = "=STDEVA(H2:H" + str(length) + ")"
            ws["K2"] = "=J2/SQRT(" + str(length-1) + ")"

            ws["M2"] = "Top 3"
            ws["M3"] = "Total"
            ws["N2"] = "=LARGE(H2:H" + str(length) + ",1)"
            ws["O2"] = "=LARGE(H2:H" + str(length) + ",2)"
            ws["P2"] = "=LARGE(H2:H" + str(length) + ",3)"
            ws["N3"] = "=SUM(N2:P2)"

        num = num+1

# remove "Sheet" worksheet, which was created automatically
del wb["Sheet"] 

# save xlsx
wb.save(path + "\\" + folder_name + ".xlsx")

# finish
print("The 1st step is finished!")
