import openpyxl
import os

# show information
print("This is version 1.0 2nd step")

# get the path of target folder
path = input("The path of target folder:")
folder_name = path.split("\\")[-1] # split with "/" and select the last one
print("folder name: " + str(folder_name))

# get the results
wb = openpyxl.load_workbook(path + "\\" + folder_name + ".xlsx", data_only = True)

name_list = []
count_list = []
top_list = []
num = 0 # num = sheet number

for sheet in wb:
    name_list.append(sheet.title)
    x = []
    for j in range(0,15): # usually 10, but for safe, I set 15
        x.append(sheet["H"+str(j+2)].value)    
    count_list.append(x)
    
    y = []
    y.append(sheet["N2"].value)
    y.append(sheet["O2"].value)
    y.append(sheet["P2"].value)
    y.append(sheet["N3"].value)
    top_list.append(y)
    
    num += 1

print("sheet number =", num)

# To let the formula stay in formula form, not value, we open the file again. But the mode is changed.
wb = openpyxl.load_workbook(path + "\\" + folder_name + ".xlsx")

# create a worksheet at the first position
wb.create_sheet("total",0)
ws = wb["total"]

# Set up the title
ws["A1"] = "Name"
ws["Q1"] = "No. 1"
ws["R1"] = "No. 2"
ws["S1"] = "No. 3"
ws["T1"] = "Total"
for i in range(0,15):
    ws[chr(i+66)+"1"] = i + 1

# import the value
for i in range(0,num):
    ws["A"+str(i+2)] = name_list[i]
    for j in range(0,15):
        ws[chr(j+66)+str(i+2)] = count_list[i][j]

    ws["Q"+str(i+2)] = top_list[i][0]
    ws["R"+str(i+2)] = top_list[i][1]
    ws["S"+str(i+2)] = top_list[i][2]
    ws["T"+str(i+2)] = top_list[i][3]

# save xlsx
wb.save(path + "\\" + folder_name + ".xlsx")

# finish
print("The 2nd step is finished!")
