import openpyxl
import os

# show informations
print("This is version 1.1 1st step")

# create a workbook
wb = openpyxl.Workbook()
wb.guess_types = True # avoid number stored in string style

# txt file number
num = 0

# get the path of target folder
path = input("The path of target folder:")
folder_name = path.split("\\")[-1] # split with "\" and select the last one
print("folder name: " + str(folder_name))

# If there is txt, read it and paste it into a new sheet
for file in os.listdir(path):
    if file.endswith(".txt"):
        print(str(num) + ". " + file)
        with open(path + "\\" + file, "r") as file_obj:
            file_name = file_obj.name.split('\\')[-1]
            name = file_name.rsplit('.',1)[0] # .rsplit('.',1)[0] to remove ".txt"
            name_number = name.split(" ")[0]
            name_cannor = name.split(" ")[1] # cannor = cancer normal

            if (name_cannor == "cancer"):
                wb.create_sheet(title = name_number) # create worksheet only when name_cannor is cancer
            
            ws = wb[name_number] # select the worksheet we created
            
            log_list = file_obj.readlines()
            length = len(log_list)
            print("cell number: " + str(length-1))

            if (name_cannor == "cancer"):
                ws["A1"] = "cancer"

            if (name_cannor == "normal"):
                ws["A19"] = "normal"
            
            for i in range(0,length):
                log_list_withoutn = log_list[i].split("\n")[0] # remove \n
                ws.append(log_list_withoutn.split("\t"))

            if (name_cannor == "cancer"):
                ws["I2"] = "average"
                ws["J2"] = "SD"
                ws["K2"] = "SEM"
                ws["I3"] = "=AVERAGE(H3:H" + str(length+1) + ")"
                ws["J3"] = "=STDEVA(H3:H" + str(length+1) + ")"
                ws["K3"] = "=J3/SQRT(" + str(length-2) + ")"

                ws["M2"] = "Top 3"
                ws["M3"] = "Total"
                ws["N2"] = "=LARGE(H3:H" + str(length+1) + ",1)"
                ws["O2"] = "=LARGE(H3:H" + str(length+1) + ",2)"
                ws["P2"] = "=LARGE(H3:H" + str(length+1) + ",3)"
                ws["N3"] = "=SUM(N2:P2)"
                
            if (name_cannor == "normal"):
                ws["I20"] = "average"
                ws["J20"] = "SD"
                ws["K20"] = "SEM"

                ws["I21"] = "=AVERAGE(H21:H" + str(length+19) + ")"
                ws["J21"] = "=STDEVA(H21:H" + str(length+19) + ")"
                ws["K21"] = "=J21/SQRT(" + str(length-2) + ")"

                ws["M20"] = "Top 3"
                ws["M21"] = "Total"
                ws["N20"] = "=LARGE(H21:H" + str(length+19) + ",1)"
                ws["O20"] = "=LARGE(H21:H" + str(length+19) + ",2)"
                ws["P20"] = "=LARGE(H21:H" + str(length+19) + ",3)"
                ws["N21"] = "=SUM(N20:P20)"

        num = num+1

# remove "Sheet" worksheet, which was created automatically
del wb["Sheet"]

# save xlsx
wb.save(path + "\\" + folder_name + ".xlsx")

# finish
print("The 1st step is finished!")
