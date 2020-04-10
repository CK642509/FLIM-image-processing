import openpyxl
import os

# show information
print("This is version 1.1 2nd step")

# get the path of target folder
path = input("The path of target folder:")
folder_name = path.split("\\")[-1] # split with "/" and select the last one
print("folder name: " + str(folder_name))

# get the results
wb = openpyxl.load_workbook(path + "\\" + folder_name + ".xlsx", data_only = True)

name_list = []
count_list_c = []
top_list_c = []
count_list_n = []
top_list_n = []
num = 0 # num = sheet number

for sheet in wb:
    name_list.append(sheet.title)

    # cancer
    count_c = []
    for j in range(0,15): # usually 10, but for safe, I set 15
        count_c.append(sheet["H"+str(j+3)].value)    
    count_list_c.append(count_c)
    
    top_c = []
    top_c.append(sheet["N2"].value)
    top_c.append(sheet["O2"].value)
    top_c.append(sheet["P2"].value)
    top_c.append(sheet["N3"].value)
    top_list_c.append(top_c)

    # normal
    count_n = []
    for j in range(0,15): # usually 10, but for safe, I set 15
        count_n.append(sheet["H"+str(j+21)].value)    
    count_list_n.append(count_n)
    
    top_n = []
    top_n.append(sheet["N20"].value)
    top_n.append(sheet["O20"].value)
    top_n.append(sheet["P20"].value)
    top_n.append(sheet["N21"].value)
    top_list_n.append(top_n)
    
    num += 1

print("sheet number =", num)

# To let the formula stay in formula form, not value, we open the file again. But the mode is changed.
wb = openpyxl.load_workbook(folder_name + ".xlsx")

wb.create_sheet("total",0) # create a worksheet at the first position
ws = wb["total"]

# Set up the title
ws["A1"] = "Name"
ws["Q1"] = "No. 1"
ws["R1"] = "No. 2"
ws["S1"] = "No. 3"
ws["T1"] = "Total"
for i in range(0,15):
    ws[chr(i+66)+"1"] = i + 1

# import the value
for i in range(0,num):
    #cancer
    ws["A"+str(i*2+2)] = name_list[i] + " cancer"
    for j in range(0,15):
        ws[chr(j+66)+str(i*2+2)] = count_list_c[i][j]

    ws["Q"+str(i*2+2)] = top_list_c[i][0]
    ws["R"+str(i*2+2)] = top_list_c[i][1]
    ws["S"+str(i*2+2)] = top_list_c[i][2]
    ws["T"+str(i*2+2)] = top_list_c[i][3]

    # normal
    ws["A"+str(i*2+3)] = name_list[i] + " normal"
    for j in range(0,15):
        ws[chr(j+66)+str(i*2+3)] = count_list_n[i][j]

    ws["Q"+str(i*2+3)] = top_list_n[i][0]
    ws["R"+str(i*2+3)] = top_list_n[i][1]
    ws["S"+str(i*2+3)] = top_list_n[i][2]
    ws["T"+str(i*2+3)] = top_list_n[i][3]

# save xlsx
wb.save(path + "\\" + folder_name + ".xlsx") 

# finish
print("The 2nd step is finished!")
